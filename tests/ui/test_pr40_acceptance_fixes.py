"""PR #40 小型整体验收修复 targeted tests（真实 GUI 行为）。

覆盖任务书第十五节 20 项：
【A 分辨率/UI】1 1920×1080 四页正常显示；2 1366×768 核心功能可访问（滚动）；
               3 1280×720 允许滚动无永久裁切；4 HistoryPage 窄窗口横向可滚动状态；
               5 SettingsPage 小窗口 settingsScrollArea 承担滚动；
               6 ProductCollection 窄窗口顶部按钮可横向访问 + 商品区纵向滚动。
【B 校准来源】7 接受 AI 重估不得污染 Excel 用户校准内容（只含用户文字）；
               8 manifest reestimate_history 仍完整；9 两次重估 → 历史页“已重估 2 次”；
               10 用户只手改 height → suggested 只含 height_cm；
               11 用户改四项 → 四项完整进入；12 程序化回填/历史恢复不产生 suggested；
               13 旧污染记录显示/导出兼容（不冒充用户建议，原始数据不破坏）。
【C AI第一次】14 历史 AI 行 / 左卡 / Excel AI首次 同一 external raw shipment。
【D tooltip】15 当前采用可编辑字段无“冻结/不可直接编辑/保守档”旧提示；
              16 真正只读自动计算字段仍保留提示。
【E Prompt】17 v2.1；18 含材质/柔软轻薄/折叠/盘绕/堆叠收拢/紧凑运输外轮廓/
              重量体积一致性/展示态≠运输态；19 无具体品类补丁；
              20 本地 PackagingEstimationService 不新增固定压缩率/类别规则。

禁止真实 API / 浏览器 / 外部链接。
"""
from __future__ import annotations

import json
import os
import sys
from dataclasses import asdict, replace

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PySide6")

from PySide6.QtCore import Qt  # noqa: E402
from PySide6.QtWidgets import QLabel, QScrollArea  # noqa: E402

import profit_accounting_26.ui.pages.calculation_page as calc_page_module  # noqa: E402
from profit_accounting_26.application import AppContext, SettingsService  # noqa: E402
from profit_accounting_26.application.calibration_export_service import (  # noqa: E402
    is_ai_reestimate_polluted_suggestion,
    user_calibration_text,
)
from profit_accounting_26.application.local_reestimate_service import LocalReestimateResult  # noqa: E402
from profit_accounting_26.application.packaging_estimation_service import PackagingEstimationService  # noqa: E402
from profit_accounting_26.application.recognition_service import RecognitionService  # noqa: E402
from profit_accounting_26.application.runtime_ai_services import (  # noqa: E402
    RecognitionOutcome,
    apply_confirmed_facts,
)
from profit_accounting_26.domain.models import (  # noqa: E402
    AIObservation,
    PackagingProposal,
    PackagingScenario,
    PackagingState,
)
from profit_accounting_26.product_collector import ProductCollectionPage  # noqa: E402
from profit_accounting_26.product_collector.collector_core.models import CandidateProduct  # noqa: E402
from profit_accounting_26.ui.pages import CalculationPage, HistoryPage, SettingsPage  # noqa: E402


# ------------------------------------------------------------------ helpers


def _scenario(label: str, *, length: float, width: float, height: float, weight: float,
              method: str = "袋装发货") -> PackagingScenario:
    return PackagingScenario(
        label=label, packaging_method=method, packaging_state=PackagingState.MODERATE_COMPRESSION,
        length_cm=length, width_cm=width, height_cm=height, weight_g=weight,
        confidence="medium", needs_review=False,
    )


def _proposal(source: str = "vision_ai_v1", *, length: float = 30.0, width: float = 20.0,
              height: float = 9.0, weight: float = 680.0) -> PackagingProposal:
    return PackagingProposal(
        normal=_scenario("AI估算", length=length, width=width, height=height, weight=weight),
        conservative=_scenario("当前采用", length=length, width=width, height=height, weight=weight),
        proposal_source=source,
        engine_version="vision-runtime-v1",
        calibration_version="",
    )


def _raw_ai_650() -> tuple[AIObservation, PackagingProposal]:
    """构造一次 AI raw 返回裸重 650g 的 V1 视觉结果（真实 parse 路径）。"""
    payload = {
        "product_name": "测试商品",
        "observed": {
            "product_price_rmb": None,
            "page_shipping_rmb": None,
            "bare_dimensions_cm": {"length": 26, "width": 16, "height": 5},
            "bare_weight_g": 650,
        },
        "bare_estimate": {"length_cm": None, "width_cm": None, "height_cm": None, "weight_g": None},
        "shipment": {"length_cm": 30, "width_cm": 20, "height_cm": 9, "weight_g": 680, "state": "袋装"},
        "note": "",
    }
    response = {"choices": [{"message": {"content": json.dumps(payload, ensure_ascii=False)}}]}
    return RecognitionService.parse_payload(response, model="vision-test")


def _outcome_700(raw_observation: AIObservation, raw_proposal: PackagingProposal) -> RecognitionOutcome:
    arbitration = AIObservation.from_dict(raw_observation.to_dict())
    conflicts = apply_confirmed_facts(
        arbitration, {"weight_g": {"value": 700, "source": "user_confirmed"}},
    )
    return RecognitionOutcome(
        raw_observation=raw_observation,
        raw_ai_proposal=raw_proposal,
        adopted_proposal=replace(raw_proposal),
        arbitration_observation=arbitration,
        arbitration_trace={"confirmed_facts_applied": {"weight_g": 700}, "conflicts": conflicts},
    )


def _reestimate_result(*, length: float, width: float, height: float, weight: float,
                       correction: str) -> LocalReestimateResult:
    raw = _proposal("corrected_reestimate_v1", length=length, width=width, height=height, weight=weight)
    adopted = _proposal("safety", length=length, width=width, height=height, weight=weight)
    return LocalReestimateResult(
        shipment=adopted.conservative,
        packaging_proposal=adopted,
        reestimate_raw_proposal=raw,
        arbitration_trace={"source": "local_reestimate_arbitration", "confirmed_facts_applied": {"weight_g": 700}},
        model="qwen3.8-max",
        provider="阿里云百炼",
    )


@pytest.fixture
def temp_context(tmp_path, monkeypatch):
    monkeypatch.setenv("PROFIT_ACCOUNTING_DATA_DIR", str(tmp_path))
    return AppContext.create_default()


@pytest.fixture
def page(qapp, temp_context):
    widget = CalculationPage(temp_context)
    yield widget
    widget.deleteLater()


def _ensure_forwarders(page):
    settings = page.context.settings_service.load()
    shenzhen = SettingsService.new_forwarder("深圳货代", 80.0, 10.0, 8000.0)
    settings["forwarders"] = [asdict(shenzhen)]
    settings["selected_forwarder_id"] = shenzhen.id
    page.context.settings_service.save(settings)
    page.refresh_settings()
    return shenzhen.id


def _silence_dialogs(monkeypatch):
    import PySide6.QtWidgets as qw

    monkeypatch.setattr(qw.QMessageBox, "information", lambda *a, **k: None)
    monkeypatch.setattr(qw.QMessageBox, "warning", lambda *a, **k: None)
    monkeypatch.setattr(calc_page_module, "confirm_action", lambda *a, **k: True)


def _arm_ai_first(page):
    """用户 700g + 第一次 AI raw 650 的完整识图状态（左卡 raw / 右卡 adopted）。"""
    page.session.confirm_value("weight_g", 700.0)
    page.bare_weight.setValue(700.0)
    raw_observation, raw_proposal = _raw_ai_650()
    page._diagnostic_operation = page.context.diagnostic_logger.begin_operation("test-round4")
    page._recognition_completed(_outcome_700(raw_observation, raw_proposal))
    return raw_observation, raw_proposal


def _accept_reestimate(page, *, length: float, width: float, height: float, weight: float,
                       correction: str) -> None:
    page._local_diagnostic_operation = page.context.diagnostic_logger.begin_operation("test-round4-re")
    page._pending_reestimate_meta = {
        "user_correction": correction,
        "confirmed_facts": {"weight_g": 700},
        "input_current_adopted": dict(page._scenario_data(page.conservative_fields)),
    }
    page._local_reestimate_completed(
        _reestimate_result(length=length, width=width, height=height, weight=weight, correction=correction)
    )


def _row_for(hp: HistoryPage, record_id: str) -> int:
    for row in range(hp.table.rowCount()):
        item = hp.table.item(row, 0)
        if item is not None and item.data(256) == record_id:
            return row
    raise AssertionError(f"record {record_id} not found in history table")


def _load_workbook(path):
    from openpyxl import load_workbook as _lw

    return _lw(str(path))


# ==================================================================
# 【A】分辨率 / 滚动行为
# ==================================================================


class TestResolutionBehavior:
    def _pages(self, temp_context):
        calc = CalculationPage(temp_context)
        hist = HistoryPage(temp_context)
        settings = SettingsPage(temp_context)
        collect = ProductCollectionPage()
        return calc, hist, settings, collect

    def test_1920_layout_without_forced_scroll(self, qapp, temp_context):
        """1. 模拟 1920×1080：四页正常显示，主布局不出现强制横向滚动。"""
        calc, hist, settings, collect = self._pages(temp_context)
        try:
            for page_widget in (calc, hist, settings, collect):
                page_widget.resize(1700, 980)
                page_widget.show()
                qapp.processEvents()
            calc_scroll = calc._root.findChild(QScrollArea, "calculationScrollArea")
            settings_scroll = settings._root.findChild(QScrollArea, "settingsScrollArea")
            assert calc_scroll.horizontalScrollBar().maximum() == 0
            assert settings_scroll.horizontalScrollBar().maximum() == 0
            assert collect._outer_scroll.horizontalScrollBar().maximum() == 0
            # 关键控件可见
            assert calc.ai_button.isVisible()
            assert collect.btn_start.isVisible()
            assert settings.btn_save.isVisible()
        finally:
            for page_widget in (calc, hist, settings, collect):
                page_widget.deleteLater()
            qapp.processEvents()

    def test_1366_and_1280_scrollable_access(self, qapp, temp_context):
        """2/3. 1366×768 与 1280×720：允许滚动访问，无功能永久裁在不可访问区域。"""
        calc, hist, settings, collect = self._pages(temp_context)
        try:
            for width, height in ((1146, 680), (1060, 630)):
                for page_widget in (calc, hist, settings, collect):
                    page_widget.resize(width, height)
                    page_widget.show()
                    qapp.processEvents()
                calc_scroll = calc._root.findChild(QScrollArea, "calculationScrollArea")
                settings_scroll = settings._root.findChild(QScrollArea, "settingsScrollArea")
                if sys.platform == "darwin":
                    # macOS 容器层整体等比缩放：窄视口不再滚动，整页按比例
                    # 缩小保证内容完整可见（用户指定的平台行为，替代滚动）
                    view = calc._scaling_view
                    scale = view.transform().m11()
                    assert calc_scroll.horizontalScrollBar().maximum() == 0
                    assert calc_scroll.verticalScrollBar().maximum() == 0
                    assert calc._root.width() * scale <= view.viewport().width() + 1
                    assert calc._root.height() * scale <= view.viewport().height() + 1
                else:
                    # 测算：body min 1360 > 窄视口 → 横向滚动可访问
                    assert calc_scroll.horizontalScrollBar().maximum() > 0
                # 设置：settingsScrollArea 承担滚动（body min 1200）
                assert (
                    settings_scroll.horizontalScrollBar().maximum() > 0
                    or settings_scroll.verticalScrollBar().maximum() > 0
                )
                # 采集：外层横向滚动可访问顶部搜索区与按钮
                assert collect._outer_scroll.horizontalScrollBar().maximum() > 0
                # 历史：等效可滚动状态（列宽总和超出视口，最右侧列可滚动访问）
                assert hist.table.horizontalHeader().length() > hist.table.viewport().width()
        finally:
            for page_widget in (calc, hist, settings, collect):
                page_widget.deleteLater()
            qapp.processEvents()

    def test_history_narrow_horizontal_scroll_state(self, qapp, temp_context):
        """4. HistoryPage 窄窗口：横向滚动策略 AsNeeded + 等效可滚动状态。"""
        hist = HistoryPage(temp_context)
        try:
            hist.resize(900, 600)
            hist.show()
            qapp.processEvents()
            assert hist.table.horizontalScrollBarPolicy() == Qt.ScrollBarPolicy.ScrollBarAsNeeded
            assert hist.table.horizontalHeader().length() > hist.table.viewport().width()
        finally:
            hist.deleteLater()
            qapp.processEvents()

    def test_settings_small_window_scroll_area(self, qapp, temp_context):
        """5. SettingsPage 小窗口：settingsScrollArea 真正承担滚动。"""
        settings = SettingsPage(temp_context)
        try:
            settings.resize(900, 600)
            settings.show()
            qapp.processEvents()
            scroll = settings._root.findChild(QScrollArea, "settingsScrollArea")
            assert (
                scroll.horizontalScrollBar().maximum() > 0
                or scroll.verticalScrollBar().maximum() > 0
            )
        finally:
            settings.deleteLater()
            qapp.processEvents()

    def test_product_collection_narrow_both_axes(self, qapp, temp_context):
        """6. ProductCollection 窄窗口：顶部按钮可横向访问；商品区纵向滚动正常。"""
        collect = ProductCollectionPage()
        try:
            collect.resize(1200, 600)
            collect.show()
            qapp.processEvents()
            # 顶部最后一个按钮（开始采集）所在整页可通过外层横向滚动访问
            assert collect._outer_scroll.horizontalScrollBar().maximum() > 0
            # 外层纵向关闭，商品卡片区继续负责纵向滚动（避免两个纵向滚动条抢滚轮）
            assert collect._outer_scroll.verticalScrollBarPolicy() == Qt.ScrollBarPolicy.ScrollBarAlwaysOff
            assert collect.scroll.verticalScrollBarPolicy() == Qt.ScrollBarPolicy.ScrollBarAsNeeded
            products = [
                CandidateProduct(
                    product_id=f"p{i}", title=f"测试商品 {i}", main_image="",
                    product_url="", keyword="测试", position=i,
                )
                for i in range(12)
            ]
            collect.load_results(products)
            qapp.processEvents()
            assert collect.scroll.verticalScrollBar().maximum() > 0
        finally:
            collect.deleteLater()
            qapp.processEvents()


# ==================================================================
# 【B】校准来源语义
# ==================================================================


class TestCalibrationProvenance:
    def test_accepted_ai_reestimate_not_in_user_calibration(self, page, monkeypatch, tmp_path):
        """7. 用户写“压缩力度再适当增大”+ 接受 AI 重估 30×22×5/750：
        Excel 用户校准内容只含用户文字，不得包含 AI 尺寸/包装说明。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        page.user_correction._widget.setPlainText("压缩力度再适当增大")
        _accept_reestimate(page, length=30, width=22, height=5, weight=750, correction="压缩力度再适当增大")
        page.recalculate()
        page.save_record()
        rid = page.record_id
        feedback = page.context.calibration_feedback_service.load(page.current_feedback_id)
        assert feedback.user_note == "压缩力度再适当增大"
        # 接受 AI 结果绝不自动变成 user_suggested
        assert feedback.suggested_package is None
        result = page.context.calibration_export_service.export(
            [page.context.store.load_record(rid)], "all", tmp_path,
        )
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        cell = workbook["校准反馈"].cell(2, 6).value  # 用户校准内容
        assert "压缩力度再适当增大" in cell
        assert "建议包装" not in cell
        assert "750" not in cell and "30×22×5" not in cell

    def test_manifest_keeps_full_reestimate_trace(self, page, monkeypatch, tmp_path):
        """8. manifest reestimate_history 仍完整包含此次 AI 重估。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        _accept_reestimate(page, length=30, width=22, height=5, weight=750, correction="压缩力度再适当增大")
        page.recalculate()
        page.save_record()
        rid = page.record_id
        v2 = page.context.history_record_v2_service.load_v2(rid)
        assert len(v2.reestimate_history) == 1
        entry = v2.reestimate_history[0]
        assert entry["adopted_reestimate_proposal"]["conservative"]["length_cm"] == 30.0
        assert entry["adopted_reestimate_proposal"]["conservative"]["weight_g"] == 750.0
        result = page.context.calibration_export_service.export(
            [page.context.store.load_record(rid)], "all", tmp_path,
        )
        manifest = json.loads((result.output_dir / "manifest.json").read_text(encoding="utf-8"))
        mf = manifest["records"][0]["machine_facts"]
        assert len(mf["reestimate_history"]) == 1
        assert mf["reestimate_history"][0]["adopted_reestimate_proposal"]["conservative"]["weight_g"] == 750

    def test_two_reestimates_history_page_shows_count(self, page, monkeypatch, qapp):
        """9. 连续接受两次重估：reestimate_history length=2、sequence=1,2，
        历史页“校准内容”显示“已重估 2 次”。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        _accept_reestimate(page, length=20, width=12, height=1, weight=35, correction="第一次修正")
        _accept_reestimate(page, length=20, width=12, height=2, weight=45, correction="第二次修正")
        page.recalculate()
        page.save_record()
        rid = page.record_id
        v2 = page.context.history_record_v2_service.load_v2(rid)
        assert len(v2.reestimate_history) == 2
        assert [entry["sequence"] for entry in v2.reestimate_history] == [1, 2]
        hp = HistoryPage(page.context)
        try:
            hp.refresh()
            qapp.processEvents()
            row = _row_for(hp, rid)
            label = hp.table.cellWidget(row, 7).findChild(QLabel)
            assert "已重估 2 次" in label.text()
        finally:
            hp.deleteLater()
            qapp.processEvents()

    def test_manual_single_field_only_in_suggested(self, page, monkeypatch):
        """10. 用户仅亲手把当前采用 height 6→5：suggested_package 只含 height_cm=5。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        page.conservative_fields["height"].spin.setValue(5.0)  # 用户亲手编辑
        page.recalculate()
        page.save_record()
        feedback = page.context.calibration_feedback_service.load(page.current_feedback_id)
        assert feedback.suggested_package is not None
        assert feedback.suggested_package.height_cm == pytest.approx(5.0)
        # 未亲手编辑的 AI 字段不得冒充用户建议
        assert feedback.suggested_package.length_cm is None
        assert feedback.suggested_package.width_cm is None
        assert feedback.suggested_package.weight_g is None

    def test_manual_all_four_fields_in_suggested(self, page, monkeypatch):
        """11. 用户亲手修改长宽高重量四项：四项完整进入 suggested_package。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        for key, value in (("length", 31.0), ("width", 21.0), ("height", 5.0), ("weight", 760.0)):
            page.conservative_fields[key].spin.setValue(value)
        page.recalculate()
        page.save_record()
        feedback = page.context.calibration_feedback_service.load(page.current_feedback_id)
        assert feedback.suggested_package is not None
        assert feedback.suggested_package.length_cm == pytest.approx(31.0)
        assert feedback.suggested_package.width_cm == pytest.approx(21.0)
        assert feedback.suggested_package.height_cm == pytest.approx(5.0)
        assert feedback.suggested_package.weight_g == pytest.approx(760.0)

    def test_programmatic_fill_and_history_restore_no_suggested(self, page, monkeypatch):
        """12. 程序化 AI 回填 / 历史恢复：不产生 suggested_package。"""
        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        _arm_ai_first(page)
        page.recalculate()
        page.save_record()
        rid = page.record_id
        # 纯 AI 复制（程序化回填）不产生校准反馈
        assert page.current_feedback_id is None
        assert page.context.calibration_feedback_service.for_record(rid) == []
        # 历史重开后再保存（无用户修改）也不产生 suggested
        page.load_record_payload(rid)
        page.save_record()
        assert page.context.calibration_feedback_service.for_record(rid) == []

    def test_legacy_polluted_suggestion_suppressed_on_display(self, temp_context, qapp):
        """13. 旧污染记录：suggested 与 AI 重估 adopted 完全一致 → 显示/导出不冒充
        用户建议；原始历史 JSON 不被破坏。"""
        from profit_accounting_26.application.calibration_export_service import user_calibration_text

        payload = {
            "id": "polluted-1",
            "product_name": "旧污染商品",
            "_v2": {
                "ai_initial": {"observation": {"product_name": "旧污染商品"}},
                "reestimate_history": [
                    {
                        "reestimate_id": "E1", "sequence": 1, "timestamp": "t", "accepted": True,
                        "adopted_reestimate_proposal": {
                            "normal": {"length_cm": 30, "width_cm": 22, "height_cm": 5, "weight_g": 750,
                                       "packaging_method": "AI重估自动包装方式"},
                            "conservative": {"length_cm": 30, "width_cm": 22, "height_cm": 5, "weight_g": 750,
                                             "packaging_method": "AI重估自动包装方式"},
                            "proposal_source": "safety",
                        },
                    },
                ],
            },
        }
        feedback_id = temp_context.calibration_feedback_service.save({
            "record_id": "polluted-1",
            "source": "user",
            "user_note": "旧记录说明",
            "suggested_package": {
                "length_cm": 30, "width_cm": 22, "height_cm": 5, "weight_g": 750,
                "packaging_method": "AI重估自动包装方式",
            },
        })
        payload["_v2"]["calibration_feedback_id"] = feedback_id
        feedback = temp_context.calibration_feedback_service.load(feedback_id)
        # 检测为旧版本污染
        assert is_ai_reestimate_polluted_suggestion(payload, feedback.suggested_package) is True
        # 导出层：不再把 AI 包装说明冒充用户校准（用户文字仍保留）
        text = user_calibration_text(feedback, payload)
        assert "用户反馈：旧记录说明" in text
        assert "建议包装" not in text
        # 原始历史 JSON 不被破坏：suggested_package 原值仍在
        raw = temp_context.calibration_feedback_service.load(feedback_id)
        assert raw.suggested_package.length_cm == 30.0
        # 历史页显示层：不把 AI 建议尺寸当作用户亲手建议
        hp = HistoryPage(temp_context)
        try:
            hp_text = hp._calibration_text(payload)
            assert "已反馈" in hp_text
            assert "750" not in hp_text and "30×22×5" not in hp_text
        finally:
            hp.deleteLater()
            qapp.processEvents()


# ==================================================================
# 【C】AI 第一次来源统一
# ==================================================================


class TestAiFirstSourceUnified:
    def test_history_left_card_excel_share_external_raw(self, page, monkeypatch, tmp_path):
        """14. 历史 AI 行 / 左侧 AI 卡 / Excel AI首次 必须来自同一 external raw。"""
        from profit_accounting_26.application.calibration_export_service import first_ai_shipment_text

        _silence_dialogs(monkeypatch)
        _ensure_forwarders(page)
        page.product_cost.setValue(66.80)
        page.domestic_shipping.setValue(28.0)
        raw_observation, raw_proposal = _raw_ai_650()  # raw shipment = 30×20×9 / 680
        adopted = _proposal("ai_candidate", length=34, width=24, height=12, weight=800)
        page._adopt_packaging(adopted)
        page.apply_proposal(page._adopted_packaging(), raw_proposal=raw_proposal)
        page._maybe_capture_initial_ai_snapshot(raw_observation, raw_proposal)
        page.recalculate()
        page.save_record()
        rid = page.record_id
        # 历史页 AI 行 = external raw（不是 adopted）
        hp = HistoryPage(page.context)
        try:
            hp.refresh()
            row = _row_for(hp, rid)
            rows = hp._packaging_rows(hp.records[row])
            assert dict(rows)["AI"] == "30×20×9 / 680g"
        finally:
            hp.deleteLater()
        # 左卡（历史重开） = external raw
        page.load_record_payload(rid)
        assert page.normal_fields["length"].value() == pytest.approx(30.0)
        assert page.normal_fields["weight"].value() == pytest.approx(680.0)
        assert page.normal_fields["length"].value() != pytest.approx(34.0)
        # Excel AI首次 = external raw
        result = page.context.calibration_export_service.export(
            [page.context.store.load_record(rid)], "all", tmp_path,
        )
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        ai_cell = workbook["校准反馈"].cell(2, 5).value  # AI首次发货估算
        assert "30×20×9" in ai_cell and "680" in ai_cell
        assert first_ai_shipment_text(page.context.store.load_record(rid)) == ai_cell


# ==================================================================
# 【D】tooltip
# ==================================================================


class TestTooltips:
    def test_current_adopted_editable_fields_no_stale_tooltip(self, page):
        """15. “当前采用”可编辑字段悬停不得出现“冻结/不可直接编辑/保守档”旧提示。"""
        for key in ("length", "width", "height", "weight"):
            spin = page.conservative_fields[key].spin
            assert spin.isReadOnly() is False
            tip = spin.toolTip()
            for stale in ("冻结", "不可直接编辑", "保守档"):
                assert stale not in tip, f"{key} tooltip 仍含旧文案: {tip}"

    def test_readonly_auto_fields_keep_correct_tooltip(self, page):
        """16. 真正只读的自动计算字段仍保留正确提示。"""
        for name in ("txtActivityPriceRmb", "txtActivityPriceUsd", "txtActivityProfitUsd"):
            spin = page._root.findChild(type(page.conservative_fields["length"].spin), name)
            assert spin is not None and spin.isReadOnly()
            assert "不可直接编辑" in spin.toolTip()
        # AI 首次发货判断只读框保留提示
        tip = page.normal_fields["method"]._widget.toolTip()
        assert "AI" in tip and "只读" in tip


# ==================================================================
# 【E】Prompt v2.1
# ==================================================================


class TestPromptV21:
    def test_version_v21(self):
        assert RecognitionService.PROMPT_VERSION == "2.6.1-visual-v2.3"

    def test_soft_thin_foldable_principle_present(self):
        """18. 全品类软/薄/可折/可盘绕运输原则。"""
        prompt = RecognitionService._prompt(1, include_json_shape=False)
        for token in (
            "材质", "柔软", "轻薄", "折叠", "盘绕", "自然堆叠", "收拢",
            "紧凑运输外轮廓", "重量与体积物理一致性", "重新检查", "展示态≠运输态",
        ):
            assert token in prompt, f"Prompt 缺少: {token}"

    def test_no_concrete_category_patch(self):
        """19. Prompt 不包含具体品类补丁。"""
        prompt = RecognitionService._prompt(1, include_json_shape=False)
        for banned in ("袜子", "绳子", "链子", "手提包", "背包", "单肩包", "包类"):
            assert banned not in prompt, f"Prompt 包含品类补丁: {banned}"

    def test_no_new_local_compression_rules(self):
        """20. 本地 PackagingEstimationService 不新增固定压缩率/类别规则。"""
        import pathlib

        import profit_accounting_26.application.packaging_estimation_service as pes_module

        source = pathlib.Path(pes_module.__file__).read_text(encoding="utf-8")
        for token in ("袜子", "手套", "绳子", "链子", "socks", "gloves"):
            assert token not in source, f"本地规则出现品类补丁: {token}"
        # 行为：完整 AI shipment（含柔软商品）无 validated 冲突时保持 AI 判断
        service = PackagingEstimationService(calibration_version="safety-test")
        observation = AIObservation(product_name="柔软轻薄商品", compressibility="good", foldability="good")
        result = service.estimate(observation, external_proposal=_proposal(weight=680.0))
        assert result.proposal_source == "ai_candidate"
        assert result.normal.weight_g == 680.0
        assert result.normal.length_cm == 30.0
